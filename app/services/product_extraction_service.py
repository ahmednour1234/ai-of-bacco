from __future__ import annotations

import asyncio
import csv
import io
import logging
import re
from functools import partial
from typing import Optional, Iterable

from fastapi import UploadFile

from app.core.exceptions import ValidationException
from app.schemas.extraction import CandidateData
from app.schemas.product_extraction import ProductExtractedItemSchema
from app.ai.normalization import QuantityParser, UnitNormalizer
from app.ai.validation import CandidateValidator

logger = logging.getLogger(__name__)

REVIEW_THRESHOLD = 0.75


class _LT:
    PRODUCT = "product"
    DESCRIPTION = "description"
    PRICE_ROW = "price_row"
    META = "meta"
    TOTAL = "total"
    IGNORE = "ignore"
    HEADER = "header"


class ProductExtractionService:
    """
    Improved extraction service.

    Goals:
    - Support PDF / image / Excel / CSV
    - Detect product rows in Arabic and English
    - Extract quantity even when it appears as a bare number inside a row
    - Handle HVAC-style rows:
        MODEL | BTU | TYPE | PRICE
        MODEL ... SAR 4,806 4 SAR 19,223
    - Handle Arabic quotation rows:
        مكيف ترين مخفي 5 طن حار وبارد انفيرتر 2 13350 26700
    - Avoid treating totals / payment / bank lines as products
    """

    def __init__(
        self,
        extra_product_keywords: Optional[set[str]] = None,
        extra_ignore_phrases: Optional[list[str]] = None,
        extra_category_keywords: Optional[dict[str, str]] = None,
        correction_examples: Optional[list[dict]] = None,
        enable_llm_qty: bool = False,
    ) -> None:
        self._extra_product_keywords: frozenset[str] = frozenset(
            kw.lower() for kw in (extra_product_keywords or set())
        )
        self._extra_ignore_patterns: tuple[re.Pattern[str], ...] = tuple(
            re.compile(p, re.IGNORECASE) for p in (extra_ignore_phrases or []) if p
        )
        self._extra_category_keywords: dict[str, str] = {
            k.lower(): v for k, v in (extra_category_keywords or {}).items()
        }
        self._correction_examples: list[dict] = correction_examples or []

        # Sub-system helpers
        self._unit_normalizer = UnitNormalizer()
        self._qty_parser = QuantityParser(enable_llm=enable_llm_qty)
        self._validator = CandidateValidator()

    # -------------------------------------------------------------------------
    # Patterns
    # -------------------------------------------------------------------------

    _QTY_UNIT_PATTERN = re.compile(
        r"(?P<qty>\d+(?:[\.,]\d+)?)\s*"
        r"(?P<unit>"
        r"kg|kgs|g|gr|gram|grams|ton|tons|طن|"
        r"pcs?|pieces?|pc|unit|units|"
        r"bag|bags|box|boxes|set|sets|carton|cartons|"
        r"ltr|liter|liters|ml|"
        r"mtr|meter|meters|"
        r"حبة|حبه|قطعة|قطعه|متر|كرتون|علبة|علبه"
        r")\b",
        re.IGNORECASE,
    )

    _BRAND_PATTERN = re.compile(
        r"(?:brand|ماركة|العلامة\s*التجارية)\s*[:\-]\s*(?P<brand>[\w\u0600-\u06FF &\./-]+)",
        re.IGNORECASE,
    )
    _CATEGORY_PATTERN = re.compile(
        r"(?:category|cat|فئة|تصنيف)\s*[:\-]\s*(?P<category>[\w\u0600-\u06FF &\./-]+)",
        re.IGNORECASE,
    )

    _ITEM_NUMBER_PREFIX = re.compile(r"^\s*(?:item\s*)?\d{1,4}[.\s)\-]+", re.IGNORECASE)

    _DIMENSION_PATTERN = re.compile(
        r'\b\d+(?:[.,]\d+)?\s*(?:mm|cm|mtr|m|ft|inch|in|ton|btu|hp|kw)\b'
        r'|\b\d+(?:[.,]\d+)?\s*(?:mm|cm|mtr|m|ft|inch|in)\s*[x×]\s*\d+(?:[.,]\d+)?\s*(?:mm|cm|mtr|m|ft|inch|in)?\b'
        r'|\b\d+(?:[.,]\d+)?\s*[x×]\s*\d+(?:[.,]\d+)?\s*(?:mm|cm|mtr|m|ft|inch|in)?\b'
        r'|\b\d+\s*/\s*\d+\s*"?'           
        r'|\b\d+(?:\.\d+)?\s*"\b'          
        r'|\b\d+(?:-\d+)?/\d+\s*"?'        
        r'|\bsch[-\s]?\d+\b'
        r'|\bdn\s*\d+\b'
        r'|\bpn\s*\d+\b'
        r'|\bdeg\b',
        re.IGNORECASE,
    )

    _PRICE_WITH_CURRENCY_RE = re.compile(
        r"(?:sar|s\.?r\.?|riyal|ريال|usd|aed|eur)\s*[:\-]?\s*(\d[\d,]*(?:\.\d+)?)",
        re.IGNORECASE,
    )

    _NUMBER_RE = re.compile(r"\d[\d,]*(?:\.\d+)?")
    _ARABIC_ONLY_RE = re.compile(r"^[\u0600-\u06FF\s\-\–\—\(\)\/]+$")
    _MODEL_CODE_RE = re.compile(r"\b[A-Z0-9]{3,}(?:[-/][A-Z0-9]+)*\b")

    _BUILTIN_PRODUCT_KEYWORDS: frozenset[str] = frozenset(
        {
            # plumbing / materials
            "pipe", "tee", "elbow", "valve", "cement", "clamp", "rod", "anchor",
            "nut", "adaptor", "adapter", "hose", "cable", "wire", "bolt", "screw",
            "bushing", "bush", "trap", "union", "wye", "reducer", "ppr", "pvc",
            "cpvc", "upvc", "hdpe", "brass", "steel", "gi", "fitting", "flange",
            "coupling", "nipple", "socket", "cap", "plug", "gasket", "washer",
            "bracket", "gauge", "strap", "pump", "motor", "switch", "conduit",
            "thread", "threaded", "flex", "gate", "ball", "check", "rubber",
            # hvac
            "cassette", "duct", "air", "outlet", "grille", "grill", "diffuser",
            "fan", "coil", "compressor", "refrigeration", "refrigerant", "cooling",
            "heating", "heat", "pump", "odU", "idu", "vrv", "tvr", "trane", "r410a",
            "outdoor", "indoor", "inverter", "ducted", "static", "pressure",
            # arabic
            "مكيف", "مضخة", "مضخه", "ماسورة", "ماسوره", "أنبوب", "انبوب", "كوع",
            "محبس", "صمام", "مخرج", "دفيوزر", "جريل", "دكت", "مروحة", "كمبرسور",
            "ضاغط", "صحراوي", "مخفي", "حار", "بارد", "انفيرتر",
        }
    )

    _BUILTIN_KEYWORDS_RE: re.Pattern[str] = re.compile(
        r"(?:"
        + "|".join(re.escape(kw) for kw in sorted(_BUILTIN_PRODUCT_KEYWORDS, key=len, reverse=True))
        + r")",
        re.IGNORECASE,
    )

    _TABLE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
        "item_no": ("item", "no", "s.no", "sr", "seq", "sl#", "sl no", "sl", "م", "الرقم", "رقم"),
        "model": ("model", "model no", "model number", "الموديل"),
        "product_name": (
            "description", "desc", "item", "product", "name", "name product",
            "اسم الصنف", "اسم المنتج", "الصنف", "البيان", "البند",
        ),
        "qty": ("qty", "quantity", "الكمية", "كمية"),
        "unit": ("unit", "uom", "الوحدة", "وحدة"),
        "price": ("price", "unit price", "rate", "سعر", "سعر الوحدة", "الافرادي", "الافرادى"),
        "total": ("amount", "total", "total amount", "الإجمالي", "الاجمالي", "المجموع"),
        "btu": ("btu",),
        "type": ("type", "refrigeration", "function", "type of function", "النوع"),
        "code": ("code", "sku", "item code", "كود الصنف", "رقم الصنف"),
        "brand": ("brand", "make", "manufacturer", "الماركة", "ماركة"),
        "category": ("category", "cat", "classification", "فئة", "تصنيف"),
    }

    _META_PATTERNS: tuple[re.Pattern[str], ...] = tuple(
        re.compile(p, re.IGNORECASE) for p in (
            r"\bquote\s*(?:no|number|#|date|expiry|ref)\b",
            r"\bquotation\b",
            r"\binvoice\b",
            r"\breference\s*[:#]",
            r"\bcustomer\s*[:#]",
            r"\bbuyer(?:'s)?\s*name\b",
            r"\bseller(?:'s)?\s*name\b",
            r"\bto\s*:",
            r"\bvat\s*(?:no|number|reg|registration)\b",
            r"\baddress\s*:",
            r"\battn\b",
            r"\bproject\s*(?:name|no|#)?\s*:",
            r"\bcontact\s*:",
            r"\bdate\s*:",
            r"\bprepared\s*by\b",
            r"\bvalid\s*(?:until|for|thru)\b",
            r"\bpage\s+\d+\s+of\s+\d+\b",
            r"رقم عرض السعر",
            r"التاريخ",
            r"اسم العميل",
            r"الرقم الضريبي",
            r"رقم السجل",
        )
    )

    _TOTAL_PATTERNS: tuple[re.Pattern[str], ...] = tuple(
        re.compile(p, re.IGNORECASE) for p in (
            r"\bgrand\s*total\b",
            r"\btotal\s+vat\b",
            r"\btotal\s+tax\b",
            r"\bdiscount\b",
            r"\bnet\s*amount\b",
            r"\bsubtotal\b",
            r"\bamount\s+in\s+words\b",
            r"\btotal\s+before\s+vat\b",
            r"\btotal\s+after\s+vat\b",
            r"\bfinal\s+price\b",
            r"\bالمجموع\s*الكلي\b",
            r"\bالاجمالي\b",
            r"\bالإجمالي\b",
            r"\bقيمة\s*الضريبة\b",
            r"\bضريبة\s*القيمة\s*المضافة\b",
            r"\bالخصم\b",
            r"\bالاجمالي\s*بعد\s*الضريبة\b",
            r"\bالإجمالي\s*بعد\s*الضريبة\b",
        )
    )

    _IGNORE_PATTERNS: tuple[re.Pattern[str], ...] = tuple(
        re.compile(p, re.IGNORECASE) for p in (
            r"\bbank\s*(?:name|account|details)\b",
            r"\baccount\s*(?:no|number)\b",
            r"\biban\b",
            r"\bswift\s*(?:code)?\b",
            r"\bpayment\s+terms?\b",
            r"\bdelivery\s*(?:method|term|time)\b",
            r"\bcurrency\s*:",
            r"\bterms?\s*(?:and|&)\s*conditions?\b",
            r"\bcomputer\s*generated\b",
            r"\bno\s+signature\b",
            r"\bsignature\s*(?:is\s*)?required\b",
            r"\bnotes?\s*:\s*",
            r"\bremarks?\s*:\s*",
            r"\bwarranty\b",
            r"\bvalidity\b",
            r"\bscope\s+of\s+work\b",
            r"\bthank you\b",
            r"\btruthfully yours\b",
            r"\bdelivery\b",
            r"\bpayment\b",
            r"\bمواد?\s+لا\s+يشملها\b",
            r"\bالضمان\b",
            r"\bطريقة\s+السداد\b",
            r"\bرقم\s+الحساب\b",
            r"\bرقم\s+الايبان\b",
            r"\bرقم\s+الايبان\b",
        )
    )

    _CATEGORY_KEYWORDS: dict[str, str] = {
        "cement": "Building Materials",
        "steel": "Building Materials",
        "paint": "Paints",
        "cable": "Electrical",
        "wire": "Electrical",
        "conduit": "Electrical",
        "pipe": "Plumbing",
        "pvc": "Plumbing",
        "ppr": "Plumbing",
        "hdpe": "Plumbing",
        "upvc": "Plumbing",
        "cpvc": "Plumbing",
        "bolt": "Hardware",
        "screw": "Hardware",
        "rod": "Hardware",
        "anchor": "Hardware",
        "nut": "Hardware",
        "valve": "Plumbing",
        "elbow": "Plumbing",
        "tee": "Plumbing",
        "fitting": "Plumbing",
        "pump": "Mechanical",
        "motor": "Mechanical",
        "clamp": "Hardware",
        "rubber": "Hardware",
        "gasket": "Hardware",
        "cassette": "HVAC",
        "duct": "HVAC",
        "diffuser": "HVAC",
        "grille": "HVAC",
        "grill": "HVAC",
        "refrigerant": "HVAC",
        "r410a": "HVAC",
        "trane": "HVAC",
        "tvr": "HVAC",
        "vrv": "HVAC",
        "مكيف": "HVAC",
        "صحراوي": "HVAC",
        "مخفي": "HVAC",
        "مضخة": "Mechanical",
        "مضخه": "Mechanical",
        "محبس": "Plumbing",
    }

    # -------------------------------------------------------------------------
    # Public API
    # -------------------------------------------------------------------------

    async def extract_candidates(self, file: UploadFile) -> list[CandidateData]:
        filename = (file.filename or "upload").strip()
        extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        file_bytes = await file.read()

        if not file_bytes:
            raise ValidationException({"file": ["Uploaded file is empty."]})

        logger.info("Extracting candidates from %r (%d bytes)", filename, len(file_bytes))

        loop = asyncio.get_event_loop()

        if extension == "csv":
            candidates = await loop.run_in_executor(
                None, self._candidates_from_csv, file_bytes
            )
        elif extension in {"xlsx", "xlsm", "xltx", "xltm"}:
            candidates = await loop.run_in_executor(
                None, self._candidates_from_excel, file_bytes
            )
        elif extension == "pdf":
            candidates = await loop.run_in_executor(
                None, self._candidates_from_pdf_bytes, file_bytes
            )
        elif extension in {"png", "jpg", "jpeg", "bmp", "tiff", "webp"}:
            text = await loop.run_in_executor(
                None, self._extract_text_from_image, file_bytes
            )
            candidates = await loop.run_in_executor(
                None, self._candidates_from_text, text
            )
        else:
            raise ValidationException(
                {"file": ["Unsupported file type. Allowed: pdf, image, excel, csv."]}
            )

        if self._correction_examples:
            await loop.run_in_executor(None, self._apply_correction_examples, candidates)

        # Normalize units and populate new fields
        await loop.run_in_executor(None, self._normalize_candidates, candidates)

        # Validate and flag suspicious values
        await loop.run_in_executor(None, self._validator.validate_candidates, candidates)

        logger.info(
            "Extracted %d candidates from %r (%d product rows)",
            len(candidates),
            filename,
            sum(1 for c in candidates if c.predicted_label == _LT.PRODUCT),
        )
        return candidates

    async def extract_from_upload(self, file: UploadFile) -> list[ProductExtractedItemSchema]:
        candidates = await self.extract_candidates(file)
        return [
            ProductExtractedItemSchema(
                product_name=c.product_name or c.raw_text[:120],
                category=c.category,
                brand=c.brand,
                quantity=c.quantity,
                unit=c.unit,
                source_line=c.raw_text,
            )
            for c in candidates
            if c.predicted_label == _LT.PRODUCT and c.product_name
        ]

    # -------------------------------------------------------------------------
    # CSV / Excel
    # -------------------------------------------------------------------------

    def _candidates_from_csv(self, file_bytes: bytes) -> list[CandidateData]:
        text = file_bytes.decode("utf-8", errors="ignore")
        return self._candidates_from_table_rows(list(csv.DictReader(io.StringIO(text))))

    def _candidates_from_excel(self, file_bytes: bytes) -> list[CandidateData]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValidationException(
                {"file": ["Excel extraction requires openpyxl package."]}
            ) from exc

        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        all_candidates: list[CandidateData] = []
        position_offset = 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)

            headers = next(rows_iter, None)
            if not headers:
                continue

            norm_headers = [self._normalize_header(str(h or "")) for h in headers]
            table_rows: list[dict[str, str]] = []
            for row in rows_iter:
                row_dict: dict[str, str] = {}
                for i, value in enumerate(row):
                    key = norm_headers[i] if i < len(norm_headers) else f"col_{i}"
                    row_dict[key] = "" if value is None else str(value)
                table_rows.append(row_dict)

            sheet_candidates = self._candidates_from_table_rows(table_rows)

            # Adjust positions so they are unique across sheets
            for c in sheet_candidates:
                c.position += position_offset
            position_offset += len(table_rows)
            all_candidates.extend(sheet_candidates)

        logger.debug("Excel: %d sheets, %d candidates total", len(workbook.sheetnames), len(all_candidates))
        return all_candidates

    def _candidates_from_table_rows(self, rows: list[dict[str, str]]) -> list[CandidateData]:
        results: list[CandidateData] = []

        for pos, row in enumerate(rows):
            product_name = self._first_non_empty(
                row,
                [
                    "product_name", "product", "item", "description", "name", "model",
                    "اسم المنتج", "اسم الصنف", "الصنف", "الوصف", "البيان",
                ],
            )
            if not product_name:
                continue

            category = self._first_non_empty(row, ["category", "cat", "classification", "تصنيف", "فئة"])
            brand = self._first_non_empty(row, ["brand", "maker", "manufacturer", "ماركة", "العلامة التجارية"])
            qty_str = self._first_non_empty(row, ["quantity", "qty", "qnty", "الكمية", "كمية"])
            raw_unit = self._first_non_empty(row, ["unit", "uom", "وحدة", "unit_of_measure"])
            price_str = self._first_non_empty(row, ["price", "unit_price", "rate", "سعر", "سعر الوحدة", "الافرادي", "الافرادى"])
            total_str = self._first_non_empty(row, ["total", "amount", "الإجمالي", "الاجمالي", "المجموع"])
            model_code = self._first_non_empty(row, ["model", "model_no", "sku", "code", "item_code", "كود"])

            quantity = self._to_float(qty_str)
            price_val = self._to_float(price_str, remove_commas=True)
            total_val = self._to_float(total_str, remove_commas=True)

            if quantity is None and price_val is not None and total_val is not None and price_val > 0:
                inferred_qty = total_val / price_val
                if self._looks_like_integerish(inferred_qty):
                    quantity = round(inferred_qty)

            # Use QuantityParser when qty_str is ambiguous or None
            if quantity is None and product_name:
                qr = self._qty_parser.parse(product_name)
                if qr.quantity is not None:
                    quantity = qr.quantity
                    if raw_unit is None:
                        raw_unit = qr.unit

            inferred_category = category or self._infer_category(product_name)
            normalized_unit = self._unit_normalizer.canonical(raw_unit)

            confidence = 0.92
            results.append(
                CandidateData(
                    raw_text=str(row),
                    predicted_label=_LT.PRODUCT,
                    confidence=confidence,
                    position=pos,
                    product_name=product_name.strip(),
                    category=inferred_category,
                    brand=brand,
                    quantity=quantity,
                    unit=normalized_unit or raw_unit,
                    price=price_val,
                    needs_review=confidence < REVIEW_THRESHOLD,
                    raw_unit=raw_unit,
                    normalized_unit=normalized_unit,
                    model_code=model_code,
                    total=total_val,
                )
            )

        return self._deduplicate_candidates(results)

    # -------------------------------------------------------------------------
    # Text pipeline
    # -------------------------------------------------------------------------

    def _normalize_candidates(self, candidates: list[CandidateData]) -> None:
        """
        Post-process in-place:
        - Normalize raw_unit → normalized_unit via UnitNormalizer
        - Use QuantityParser to fill missing quantity from raw_text when needed
        - Extract model_code via regex when not already set
        - Canonicalize the unit field to normalized_unit (fallback raw_unit)
        """
        for cand in candidates:
            if cand.predicted_label != _LT.PRODUCT:
                continue

            # Normalize unit
            if cand.raw_unit is None and cand.unit:
                cand.raw_unit = cand.unit
            normalized = self._unit_normalizer.canonical(cand.raw_unit or cand.unit)
            if normalized:
                cand.normalized_unit = normalized
                cand.unit = normalized

            # Fill missing quantity from raw_text via QuantityParser
            if cand.quantity is None and cand.raw_text:
                qr = self._qty_parser.parse(cand.raw_text)
                if qr.quantity is not None:
                    cand.quantity = qr.quantity
                    if cand.raw_unit is None and qr.unit:
                        cand.raw_unit = qr.unit
                        cand.normalized_unit = self._unit_normalizer.canonical(qr.unit) or qr.unit
                        cand.unit = cand.normalized_unit

            # Extract model code from raw_text if not set
            if cand.model_code is None and cand.raw_text:
                m = self._MODEL_CODE_RE.search(cand.raw_text)
                if m:
                    cand.model_code = m.group(0)

    def _candidates_from_text(self, text: str) -> list[CandidateData]:
        lines = self._normalize_text_lines(text)

        # 1) structured rows with headers
        table_results = self._try_table_extraction(lines)
        if table_results:
            return table_results

        # 2) paired product line + numeric line
        paired_results = self._try_product_numeric_pair_extraction(lines)
        if paired_results:
            return paired_results

        # 3) fallback sequential classification
        return self._sequential_line_extraction(lines)

    def _try_table_extraction(self, lines: list[str]) -> Optional[list[CandidateData]]:
        header_idx: Optional[int] = None
        header_map: Optional[dict[str, int]] = None

        for idx, line in enumerate(lines):
            mapped = self._map_header_columns(line)
            if mapped is not None:
                header_idx = idx
                header_map = mapped
                break

        if header_idx is None or header_map is None:
            return None

        results: list[CandidateData] = []
        data_lines = lines[header_idx + 1:]

        for pos, line in enumerate(data_lines):
            if self._looks_like_meta_line(line) or self._looks_like_total_line(line) or self._looks_like_ignore_line(line):
                break

            cells = self._split_cells(line)
            if not cells:
                continue

            row_dict = self._cells_to_dict(cells, header_map)

            product_name = self._first_non_empty(
                row_dict,
                ["product_name", "description", "desc", "item", "product", "model", "name", "code"],
            )
            if not product_name:
                continue

            qty_str = self._first_non_empty(row_dict, ["qty", "quantity"])
            unit = self._first_non_empty(row_dict, ["unit", "uom"])
            price_str = self._first_non_empty(row_dict, ["price", "rate", "unit_price"])
            total_str = self._first_non_empty(row_dict, ["amount", "total"])
            brand = self._first_non_empty(row_dict, ["brand", "make"])
            category = self._first_non_empty(row_dict, ["category", "cat"])

            quantity = self._to_float(qty_str, remove_commas=True)
            price_val = self._to_float(price_str, remove_commas=True)
            total_val = self._to_float(total_str, remove_commas=True)

            if quantity is None:
                quantity = self._extract_qty_from_inline_row(line)

            if quantity is None and price_val is not None and total_val is not None and price_val > 0:
                inferred_qty = total_val / price_val
                if self._looks_like_integerish(inferred_qty):
                    quantity = round(inferred_qty)

            inferred_category = category or self._infer_category(product_name)
            raw_unit = unit
            normalized_unit = self._unit_normalizer.canonical(raw_unit)

            # Extract model code from product_name or code field
            model_code_str: Optional[str] = self._first_non_empty(row_dict, ["code", "sku", "model"])
            if not model_code_str:
                m_code = self._MODEL_CODE_RE.search(product_name)
                model_code_str = m_code.group(0) if m_code else None

            # Continuation line: product description only, no numeric data.
            # Merge into the previous product's name instead of creating a new entry.
            if quantity is None and price_val is None and total_val is None:
                if results:
                    results[-1].product_name = (
                        (results[-1].product_name or "") + " " + product_name.strip()
                    ).strip()
                continue

            confidence = 0.90
            results.append(
                CandidateData(
                    raw_text=line,
                    predicted_label=_LT.PRODUCT,
                    confidence=confidence,
                    position=pos,
                    product_name=product_name.strip(),
                    category=inferred_category,
                    brand=brand,
                    quantity=quantity,
                    unit=normalized_unit or raw_unit,
                    price=price_val,
                    needs_review=confidence < REVIEW_THRESHOLD,
                    raw_unit=raw_unit,
                    normalized_unit=normalized_unit,
                    model_code=model_code_str,
                    total=total_val,
                )
            )

        return self._deduplicate_candidates(results) if results else None

    def _try_product_numeric_pair_extraction(self, lines: list[str]) -> Optional[list[CandidateData]]:
        """
        Handles documents like:
        PRODUCT NAME...
        3.00 4,630.00 13,890.00

        and Arabic:
        مكيف ترين مخفي 5 طن حار وبارد انفيرتر 2 13350 26700
        """
        results: list[CandidateData] = []
        i = 0
        pos = 0

        while i < len(lines):
            line = lines[i]

            if self._looks_like_meta_line(line) or self._looks_like_total_line(line) or self._looks_like_ignore_line(line):
                i += 1
                continue

            # single-row product with inline numbers
            if self._looks_like_inline_product_row(line):
                product_name, quantity, price_val = self._parse_inline_product_row(line)
                if product_name:
                    results.append(
                        CandidateData(
                            raw_text=line,
                            predicted_label=_LT.PRODUCT,
                            confidence=0.88,
                            position=pos,
                            product_name=product_name,
                            category=self._infer_category(product_name),
                            brand=None,
                            quantity=quantity,
                            unit=None,
                            price=price_val,
                            needs_review=False,
                        )
                    )
                    pos += 1
                    i += 1
                    continue

            if not self._looks_like_product_line(line):
                i += 1
                continue

            product_parts = [line]
            j = i + 1

            while j < len(lines):
                next_line = lines[j]

                if self._looks_like_meta_line(next_line) or self._looks_like_total_line(next_line) or self._looks_like_ignore_line(next_line):
                    break

                if self._looks_like_numeric_row(next_line):
                    full_name = self._merge_text_parts(product_parts)
                    quantity = self._extract_qty_from_numeric_row(next_line)
                    price_val = self._extract_unit_price_from_numeric_row(next_line)
                    results.append(
                        CandidateData(
                            raw_text=full_name + " || " + next_line,
                            predicted_label=_LT.PRODUCT,
                            confidence=0.89,
                            position=pos,
                            product_name=full_name,
                            category=self._infer_category(full_name),
                            brand=None,
                            quantity=quantity,
                            unit=None,
                            price=price_val,
                            needs_review=False,
                        )
                    )
                    pos += 1
                    i = j + 1
                    break

                if self._looks_like_continuation_line(next_line):
                    product_parts.append(next_line)
                    j += 1
                    continue

                break
            else:
                i += 1
                continue

            if i < j + 1:
                continue

            i += 1

        return self._deduplicate_candidates(results) if results else None

    def _sequential_line_extraction(self, lines: list[str]) -> list[CandidateData]:
        results: list[CandidateData] = []
        pending: Optional[CandidateData] = None
        pos = 0

        for line in lines:
            label, score = self._classify_line(line)

            if label == _LT.PRODUCT:
                if pending is not None:
                    results.append(pending)
                confidence = self._score_to_confidence(score)
                pending = self._build_product_candidate(line, pos, confidence)
                pos += 1
                continue

            if label == _LT.DESCRIPTION:
                if pending is not None:
                    pending.description = self._merge_text(pending.description, line)
                continue

            if label == _LT.PRICE_ROW:
                if pending is not None:
                    if pending.quantity is None:
                        pending.quantity = self._extract_qty_from_numeric_row(line) or self._extract_qty_from_inline_row(line)
                    if pending.unit is None:
                        _, unit = self._extract_qty_unit(line)
                        pending.unit = unit
                    if pending.price is None:
                        pending.price = self._extract_unit_price_from_numeric_row(line) or self._extract_price(line)
                continue

            if label in {_LT.META, _LT.TOTAL, _LT.IGNORE, _LT.HEADER}:
                if pending is not None:
                    results.append(pending)
                    pending = None
                continue

        if pending is not None:
            results.append(pending)

        return self._deduplicate_candidates(results)

    # -------------------------------------------------------------------------
    # Classification
    # -------------------------------------------------------------------------

    def _classify_line(self, line: str) -> tuple[str, int]:
        if self._looks_like_ignore_line(line):
            return _LT.IGNORE, 0
        if self._looks_like_meta_line(line):
            return _LT.META, 0
        if self._looks_like_total_line(line):
            return _LT.TOTAL, 0
        if self._map_header_columns(line) is not None:
            return _LT.HEADER, 0

        score = self._product_score(line)
        if score >= 3:
            return _LT.PRODUCT, score

        if self._looks_like_numeric_row(line):
            return _LT.PRICE_ROW, 0

        if self._looks_like_description_line(line):
            return _LT.DESCRIPTION, 0

        return _LT.IGNORE, 0

    def _product_score(self, line: str) -> int:
        score = 0

        if self._ITEM_NUMBER_PREFIX.match(line):
            rest = self._ITEM_NUMBER_PREFIX.sub("", line).strip()
            if len(rest) > 4 and not rest.isdigit():
                score += 3

        kw_matches = set(m.lower() for m in self._BUILTIN_KEYWORDS_RE.findall(line))
        if len(kw_matches) >= 2:
            score += 3
        elif len(kw_matches) == 1:
            score += 1

        if self._DIMENSION_PATTERN.search(line):
            score += 2

        if self._MODEL_CODE_RE.search(line):
            score += 1

        lower_line = line.lower()
        for kw in self._extra_product_keywords:
            if re.search(r"\b" + re.escape(kw) + r"\b", lower_line):
                score += 2

        return score

    def _looks_like_product_line(self, line: str) -> bool:
        return self._product_score(line) >= 3

    def _looks_like_description_line(self, line: str) -> bool:
        if len(line) > 160:
            return False
        if self._looks_like_numeric_row(line):
            return False
        if self._ARABIC_ONLY_RE.match(line):
            return True
        arabic_chars = sum(1 for c in line if "\u0600" <= c <= "\u06FF")
        return arabic_chars / max(len(line), 1) > 0.35

    def _looks_like_numeric_row(self, line: str) -> bool:
        if self._looks_like_total_line(line):
            return False
        numbers = self._extract_numbers(line)
        if len(numbers) < 2:
            return False
        if self._QTY_UNIT_PATTERN.search(line):
            return True

        stripped = re.sub(r"(sar|s\.?r\.?|ريال|usd|aed|eur|%|\(|\)|,|\.)", " ", line, flags=re.IGNORECASE)
        alpha = re.findall(r"[A-Za-z\u0600-\u06FF]+", stripped)
        return len(alpha) <= 2 and len(numbers) >= 2

    def _looks_like_inline_product_row(self, line: str) -> bool:
        if self._looks_like_meta_line(line) or self._looks_like_total_line(line) or self._looks_like_ignore_line(line):
            return False
        if not self._looks_like_product_line(line):
            return False

        numbers = self._extract_numbers(line)
        if len(numbers) < 3:
            return False

        return True

    def _looks_like_continuation_line(self, line: str) -> bool:
        if self._looks_like_meta_line(line) or self._looks_like_total_line(line) or self._looks_like_ignore_line(line):
            return False
        if self._looks_like_numeric_row(line):
            return False

        # usually continuation line has text and few/no numbers
        alpha_count = len(re.findall(r"[A-Za-z\u0600-\u06FF]+", line))
        return alpha_count >= 1

    def _looks_like_meta_line(self, line: str) -> bool:
        return any(p.search(line) for p in self._META_PATTERNS)

    def _looks_like_total_line(self, line: str) -> bool:
        return any(p.search(line) for p in self._TOTAL_PATTERNS)

    def _looks_like_ignore_line(self, line: str) -> bool:
        if any(p.search(line) for p in self._IGNORE_PATTERNS):
            return True
        if any(p.search(line) for p in self._extra_ignore_patterns):
            return True
        return False

    # -------------------------------------------------------------------------
    # Table helpers
    # -------------------------------------------------------------------------

    def _map_header_columns(self, line: str) -> Optional[dict[str, int]]:
        cells = self._split_cells(line)
        if not cells:
            return None

        mapped: dict[str, int] = {}
        for idx, cell in enumerate(cells):
            norm = self._normalize_header(cell)
            canonical = self._resolve_header_alias(norm)
            if canonical:
                mapped[canonical] = idx

        # must detect product + at least one pricing/qty context
        if "product_name" in mapped and ({"qty", "price", "total", "model"} & set(mapped.keys())):
            return mapped

        # MODEL / BTU / TYPE / PRICE image
        if "model" in mapped and "price" in mapped:
            return mapped

        # Arabic header row
        if "product_name" in mapped and "qty" in mapped:
            return mapped

        return None

    def _resolve_header_alias(self, normalized_header: str) -> Optional[str]:
        # First pass: exact match (prevents "unit" matching "unit price")
        for canonical, aliases in self._TABLE_HEADER_ALIASES.items():
            for alias in aliases:
                alias_norm = self._normalize_header(alias)
                if normalized_header == alias_norm:
                    return canonical
        # Second pass: substring match
        for canonical, aliases in self._TABLE_HEADER_ALIASES.items():
            for alias in aliases:
                alias_norm = self._normalize_header(alias)
                if alias_norm in normalized_header or normalized_header in alias_norm:
                    return canonical
        return None

    def _split_cells(self, line: str) -> list[str]:
        cells = [c.strip() for c in re.split(r"\s{2,}|\t|[|]", line) if c.strip()]
        if len(cells) > 1:
            return cells

        # fallback for OCR lines with single spaces only
        return [line.strip()] if line.strip() else []

    def _cells_to_dict(self, cells: list[str], header_map: dict[str, int]) -> dict[str, str]:
        row: dict[str, str] = {}
        for key, index in header_map.items():
            val = cells[index] if index < len(cells) else ""
            row[key] = "" if val == "-" else val
        return row

    # -------------------------------------------------------------------------
    # Parsing product rows
    # -------------------------------------------------------------------------

    def _parse_inline_product_row(self, line: str) -> Optional[tuple[str]Optional[, float]Optional[, float]]:
        """
        Example:
        مكيف ترين مخفي 5 طن حار وبارد انفيرتر 2 13350 26700
        -> name, qty=2, price=13350

        Example:
        4TVDD055AB07WAA ... SAR 4,806 4 SAR 19,223
        -> name, qty=4, price=4806
        """
        quantity = self._extract_qty_from_inline_row(line)
        price_val = self._extract_unit_price_from_inline_row(line)
        product_name = self._extract_product_name_from_inline_row(line)

        return product_name, quantity, price_val

    def _extract_product_name_from_inline_row(self, line: str) -> Optional[str]:
        clean = self._ITEM_NUMBER_PREFIX.sub("", line).strip()

        # SAR price qty SAR total
        m = re.search(
            r"^(?P<name>.+?)\s+SAR\s+\d[\d,]*(?:\.\d+)?\s+\d+(?:\.\d+)?\s+SAR\s+\d[\d,]*(?:\.\d+)?\s*$",
            clean,
            re.IGNORECASE,
        )
        if m:
            return m.group("name").strip()

        # generic: name + qty + price + total
        tokens = clean.split()
        if len(tokens) >= 4:
            tail = tokens[-3:]
            if all(self._is_numeric_token(t) for t in tail):
                return " ".join(tokens[:-3]).strip()

        return clean if self._looks_like_product_line(clean) else None

    def _extract_qty_from_inline_row(self, line: str) -> Optional[float]:
        # explicit qty + unit first
        qty, _ = self._extract_qty_unit(line)
        if qty is not None:
            return qty

        # pattern: SAR 4,806 4 SAR 19,223 -> qty=4
        m = re.search(
            r"SAR\s+\d[\d,]*(?:\.\d+)?\s+(?P<qty>\d+(?:\.\d+)?)\s+SAR\s+\d[\d,]*(?:\.\d+)?",
            line,
            re.IGNORECASE,
        )
        if m:
            return self._to_float(m.group("qty"), remove_commas=True)

        nums = self._extract_numbers(line)
        if len(nums) < 3:
            return None

        # Arabic price list: name qty unit_price total
        maybe_qty = nums[-3]
        maybe_price = nums[-2]
        maybe_total = nums[-1]

        if maybe_price > 0 and maybe_total >= maybe_price:
            ratio = maybe_total / maybe_price
            if self._looks_like_integerish(ratio) and abs(ratio - maybe_qty) < 0.01:
                return maybe_qty

        # fallback: if third-from-last is small integer
        if 0 < maybe_qty <= 1000 and self._looks_like_integerish(maybe_qty):
            return maybe_qty

        return None

    def _extract_unit_price_from_inline_row(self, line: str) -> Optional[float]:
        # SAR unit price qty SAR total
        m = re.search(
            r"SAR\s+(?P<price>\d[\d,]*(?:\.\d+)?)\s+\d+(?:\.\d+)?\s+SAR\s+\d[\d,]*(?:\.\d+)?",
            line,
            re.IGNORECASE,
        )
        if m:
            return self._to_float(m.group("price"), remove_commas=True)

        nums = self._extract_numbers(line)
        if len(nums) >= 3:
            return nums[-2]
        return self._extract_price(line)

    def _extract_qty_from_numeric_row(self, line: str) -> Optional[float]:
        qty, _ = self._extract_qty_unit(line)
        if qty is not None:
            return qty

        nums = self._extract_numbers(line)
        if len(nums) < 2:
            return None

        # common patterns:
        # 3.00 4,630.00 13,890.00 -> qty price total
        if len(nums) >= 3:
            a, b, c = nums[0], nums[1], nums[2]
            if b > 0 and c >= b:
                ratio = c / b
                if self._looks_like_integerish(ratio) and abs(ratio - a) < 0.01:
                    return a

        # fallback first number if small
        first = nums[0]
        if 0 < first <= 1000:
            return first

        return None

    def _extract_unit_price_from_numeric_row(self, line: str) -> Optional[float]:
        nums = self._extract_numbers(line)
        if len(nums) >= 2:
            # qty, unit_price, total
            return nums[1]
        return self._extract_price(line)

    # -------------------------------------------------------------------------
    # Candidate builder
    # -------------------------------------------------------------------------

    def _build_product_candidate(self, line: str, position: int, confidence: float) -> CandidateData:
        clean = self._ITEM_NUMBER_PREFIX.sub("", line).strip()

        brand: Optional[str] = None
        brand_match = self._BRAND_PATTERN.search(clean)
        if brand_match:
            brand = brand_match.group("brand").strip()
            clean = self._BRAND_PATTERN.sub("", clean).strip()

        category: Optional[str] = None
        category_match = self._CATEGORY_PATTERN.search(clean)
        if category_match:
            category = category_match.group("category").strip()
            clean = self._CATEGORY_PATTERN.sub("", clean).strip()

        product_name = re.sub(r"\s+", " ", clean).strip(" -|,;:")
        if not category:
            category = self._infer_category(product_name)

        return CandidateData(
            raw_text=line,
            predicted_label=_LT.PRODUCT,
            confidence=confidence,
            position=position,
            product_name=product_name,
            category=category,
            brand=brand,
            quantity=None,
            unit=None,
            price=None,
            needs_review=confidence < REVIEW_THRESHOLD,
        )

    # -------------------------------------------------------------------------
    # Correction examples
    # -------------------------------------------------------------------------

    def _apply_correction_examples(self, candidates: list[CandidateData]) -> None:
        if not self._correction_examples:
            return

        for cand in candidates:
            query_words = set(re.findall(r"\b\w{3,}\b", cand.raw_text.lower()))
            if not query_words:
                continue

            best_score = 0.0
            best_example: Optional[dict] = None

            for ex in self._correction_examples:
                ex_words = set(re.findall(r"\b\w{3,}\b", ex.get("normalized_text", "").lower()))
                if not ex_words:
                    continue
                union = query_words | ex_words
                inter = query_words & ex_words
                sim = len(inter) / len(union) if union else 0.0
                if sim > best_score:
                    best_score = sim
                    best_example = ex

            if best_score >= 0.60 and best_example is not None:
                cand.predicted_label = best_example.get("correct_label", cand.predicted_label)
                cand.confidence = 0.95
                cand.needs_review = False
                if best_example.get("correct_name"):
                    cand.product_name = best_example["correct_name"]
                if best_example.get("correct_category"):
                    cand.category = best_example["correct_category"]
                if best_example.get("correct_brand"):
                    cand.brand = best_example["correct_brand"]

    # -------------------------------------------------------------------------
    # Readers
    # -------------------------------------------------------------------------

    def _extract_text_from_pdf(self, file_bytes: bytes) -> str:
        """Extract plain text from a PDF (used as fallback in _candidates_from_pdf_bytes)."""
        try:
            import pdfplumber  # type: ignore
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                pages_text: list[str] = []
                for page in pdf.pages:
                    pages_text.append(page.extract_text() or "")
                text = "\n".join(pages_text)
                if text.strip():
                    logger.debug("PDF: extracted %d chars via pdfplumber", len(text))
                    return text
        except Exception as plumber_err:
            logger.debug("pdfplumber plain-text failed (%s), using pypdf", plumber_err)

        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise ValidationException({"file": ["PDF extraction requires pypdf package."]}) from exc

        try:
            reader = PdfReader(io.BytesIO(file_bytes), strict=False)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            logger.debug("PDF: extracted %d chars via pypdf", len(text))
            return text
        except Exception as pdf_err:
            logger.warning("pypdf extraction failed: %s", pdf_err)
            raise ValidationException(
                {"file": [f"Could not read PDF: it may be encrypted, corrupted, or image-only. ({type(pdf_err).__name__})"]}
            ) from pdf_err

    def _candidates_from_pdf_bytes(self, file_bytes: bytes) -> list[CandidateData]:
        """
        Unified PDF extraction pipeline:
        1. pdfplumber word-coordinate extraction → column-aware table parsing
        2. fallback: plain-text → _candidates_from_text
        """
        structured_lines: list[str] = []
        try:
            import pdfplumber  # type: ignore
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    words = page.extract_words(
                        x_tolerance=5, y_tolerance=3, keep_blank_chars=False
                    )
                    if words:
                        structured_lines.extend(self._words_to_table_lines(words))
                    else:
                        structured_lines.extend(
                            (page.extract_text() or "").splitlines()
                        )
        except Exception as plumber_err:
            logger.debug(
                "pdfplumber word extraction failed (%s), falling back to pypdf",
                plumber_err,
            )

        if structured_lines:
            # Pass structured lines directly to table extractor — do NOT call
            # _normalize_text_lines which strips pipe chars and corrupts column format.
            table_results = self._try_table_extraction(structured_lines)
            if table_results:
                logger.debug(
                    "PDF: %d product(s) via coordinate-based extraction",
                    len(table_results),
                )
                return table_results
            # Table detection failed; fall back to text-based candidate extraction
            text = "\n".join(structured_lines)
            fallback = self._candidates_from_text(text)
            if fallback:
                return fallback

        # pypdf plain-text fallback
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise ValidationException(
                {"file": ["PDF extraction requires pypdf package."]}
            ) from exc
        try:
            reader = PdfReader(io.BytesIO(file_bytes), strict=False)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            logger.debug("PDF: fallback to pypdf, %d chars", len(text))
            return self._candidates_from_text(text)
        except Exception as pdf_err:
            logger.warning("pypdf extraction failed: %s", pdf_err)
            raise ValidationException(
                {"file": [f"Could not read PDF: it may be encrypted, corrupted, or image-only. ({type(pdf_err).__name__})"]}
            ) from pdf_err

    @staticmethod
    def _words_to_table_lines(words: list[dict]) -> list[str]:
        """
        Reconstruct table lines from pdfplumber word dicts with column-position awareness.

        Algorithm:
        1. Group words into Y-rows (4 pt bucket tolerance).
        2. Find the header row — row with the most column-keyword matches.
        3. Merge adjacent header words within HEADER_MERGE_GAP pt into one cell;
           each merged cell's x0 becomes a column anchor.
        4. For every row, assign each word to the rightmost anchor whose x0 <= word.x0 + ANCHOR_SLACK.
        5. Output one pipe-separated string per row; "-" fills empty column slots.

        The fixed column count matches the header, keeping positional indices
        consistent for _cells_to_dict (critical for _try_table_extraction).
        """
        HEADER_MERGE_GAP: float = 5.0    # pt: merge tightly adjacent words in header
        ANCHOR_SLACK: float = 15.0       # pt: right-aligned values may sit left of anchor
        HEADER_KEYWORDS = frozenset({
            "product", "description", "desc", "item", "name",
            "qty", "quantity", "unit", "uom", "price", "rate",
            "amount", "total", "sl#", "discount", "vat", "tax",
            "\u0627\u0633\u0645", "\u0643\u0645\u064a\u0629",
            "\u0633\u0639\u0631", "\u0627\u0644\u0648\u062d\u062f\u0629",
            "\u0627\u062c\u0645\u0627\u0644\u064a",
        })

        # ── 1. group words into Y-bucket rows ─────────────────────────────
        rows_map: dict[int, list[dict]] = {}
        for w in words:
            y_key = round(w["top"] / 4) * 4
            rows_map.setdefault(y_key, []).append(w)
        sorted_rows = [
            (y, sorted(ws, key=lambda w: w["x0"]))
            for y, ws in sorted(rows_map.items())
        ]
        if not sorted_rows:
            return []

        # ── 2. find best-matching header row ──────────────────────────────
        page_height = max(w["top"] for w in words)
        min_y = page_height * 0.12   # skip logo / address block at top
        best_score = 0
        header_row_idx = 0
        for i, (y, ws) in enumerate(sorted_rows):
            if y < min_y:
                continue
            row_text = " ".join(w["text"].lower() for w in ws)
            score = sum(1 for kw in HEADER_KEYWORDS if kw in row_text)
            if score > best_score:
                best_score = score
                header_row_idx = i

        if best_score < 2:
            # No structured table — return plain row text
            return [" ".join(w["text"] for w in ws) for _, ws in sorted_rows]

        _, header_words = sorted_rows[header_row_idx]

        # ── 3. build column anchors by merging adjacent header words ──────
        col_anchors: list[float] = []
        cur_x0 = header_words[0]["x0"]
        cur_tokens: list[str] = [header_words[0]["text"]]
        prev_x1: float = header_words[0]["x1"]
        for w in header_words[1:]:
            gap = w["x0"] - prev_x1
            if gap < HEADER_MERGE_GAP:
                cur_tokens.append(w["text"])
            else:
                col_anchors.append(cur_x0)
                cur_x0 = w["x0"]
                cur_tokens = [w["text"]]
            prev_x1 = w["x1"]
        col_anchors.append(cur_x0)
        n_cols = len(col_anchors)

        def assign_col(x0: float) -> int:
            col = 0
            for i in range(n_cols):
                if col_anchors[i] <= x0 + ANCHOR_SLACK:
                    col = i
            return col

        # ── 4. build output lines ──────────────────────────────────────────
        lines: list[str] = []
        for _y, row_words in sorted_rows:
            if not row_words:
                continue
            cols: dict[int, list[str]] = {}
            for w in row_words:
                c = assign_col(w["x0"])
                cols.setdefault(c, []).append(w["text"])
            cells = [
                " ".join(cols[i]) if i in cols else "-"
                for i in range(n_cols)
            ]
            lines.append("|".join(cells))
        return lines

    def _extract_text_from_image(self, file_bytes: bytes) -> str:
        try:
            from PIL import Image
            import pytesseract
        except ImportError as exc:
            raise ValidationException(
                {"file": ["Image extraction requires Pillow and pytesseract packages."]}
            ) from exc

        image = Image.open(io.BytesIO(file_bytes))
        return pytesseract.image_to_string(image)

    # -------------------------------------------------------------------------
    # Utils
    # -------------------------------------------------------------------------

    @staticmethod
    def _score_to_confidence(score: int) -> float:
        table = {0: 0.00, 1: 0.30, 2: 0.55, 3: 0.75, 4: 0.85}
        if score in table:
            return table[score]
        return min(0.99, 0.85 + 0.02 * (score - 4))

    def _normalize_text_lines(self, text: str) -> list[str]:
        lines: list[str] = []
        for raw in text.splitlines():
            line = re.sub(r"\s+", " ", raw).strip()
            line = line.strip("|-_")
            if len(line) >= 2:
                lines.append(line)
        return lines

    def _extract_qty_unit(self, line: str) -> Optional[tuple[float]Optional[, str]]:
        match = self._QTY_UNIT_PATTERN.search(line)
        if not match:
            return None, None
        qty = self._to_float(match.group("qty"))
        unit = match.group("unit")
        return qty, unit

    def _extract_price(self, line: str) -> Optional[float]:
        m = self._PRICE_WITH_CURRENCY_RE.search(line)
        if m:
            return self._to_float(m.group(1), remove_commas=True)

        nums = self._extract_numbers(line)
        if nums:
            return nums[-1]
        return None

    def _extract_numbers(self, text: str) -> list[float]:
        values: list[float] = []
        for token in self._NUMBER_RE.findall(text):
            value = self._to_float(token, remove_commas=True)
            if value is not None:
                values.append(value)
        return values

    def _infer_category(self, name: str) -> Optional[str]:
        lower = name.lower()
        for keyword, category in self._extra_category_keywords.items():
            if keyword in lower:
                return category
        for keyword, category in self._CATEGORY_KEYWORDS.items():
            if keyword in lower:
                return category
        return None

    def _normalize_header(self, header: str) -> str:
        return re.sub(r"\s+", "_", header.strip().lower())

    def _first_non_empty(self, row: dict[str, str], keys: Iterable[str]) -> Optional[str]:
        normalized_row = {self._normalize_header(k): v for k, v in row.items()}
        for key in keys:
            value = normalized_row.get(self._normalize_header(key))
            if value is None:
                continue
            candidate = str(value).strip()
            if candidate:
                return candidate
        return None

    def _merge_text(self, first: Optional[str], second: Optional[str]) -> Optional[str]:
        if first and second:
            return f"{first} {second}".strip()
        return first or second

    def _merge_text_parts(self, parts: list[str]) -> str:
        return re.sub(r"\s+", " ", " ".join(p.strip() for p in parts if p.strip())).strip()

    def _deduplicate_candidates(self, candidates: list[CandidateData]) -> list[CandidateData]:
        seen: set[str] = set()
        unique: list[CandidateData] = []

        for c in candidates:
            key = "|".join(
                [
                    (c.product_name or "").strip().lower(),
                    "" if c.quantity is None else str(c.quantity),
                    "" if c.price is None else str(c.price),
                ]
            )
            if key in seen:
                continue
            seen.add(key)
            unique.append(c)

        return unique

    @staticmethod
    def _to_float(value: Optional[str], remove_commas: bool = False) -> Optional[float]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        if remove_commas:
            text = text.replace(",", "")
        else:
            # if there is one comma and no dot, treat comma as decimal
            if "," in text and "." not in text and text.count(",") == 1:
                text = text.replace(",", ".")
            else:
                text = text.replace(",", "")
        try:
            return float(text)
        except ValueError:
            return None

    @staticmethod
    def _is_numeric_token(token: str) -> bool:
        return bool(re.fullmatch(r"\d[\d,]*(?:\.\d+)?", token))

    @staticmethod
    def _looks_like_integerish(value: float, tolerance: float = 0.05) -> bool:
        return abs(value - round(value)) <= tolerance
