"""
app/ai/parsers/excel_parser.py
==============================
Excel (.xlsx, .xls, .xlsm) → DocumentRepresentation

Each sheet becomes a DocumentTable.
Header detection: first non-empty row whose values are predominantly non-numeric.
All sheets are enumerated; the flat all_rows list merges all sheets in order.
"""

from __future__ import annotations

import io
import re
import unicodedata

from app.ai.interfaces.base_parser import BaseParser
from app.schemas.document_representation import (
    DocumentBlock,
    DocumentRepresentation,
    DocumentTable,
)

try:
    import openpyxl  # type: ignore
    from openpyxl.utils.exceptions import InvalidFileException  # type: ignore
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

_ARABIC_RE = re.compile(r"[\u0600-\u06FF\u0750-\u077F\uFB50-\uFDFF\uFE70-\uFEFF]")


def _detect_language(text: str) -> tuple[str, bool]:
    if not text:
        return "unknown", False
    arabic_chars = len(_ARABIC_RE.findall(text))
    latin_chars = sum(1 for c in text if c.isascii() and c.isalpha())
    total = arabic_chars + latin_chars
    if total == 0:
        return "unknown", False
    ratio = arabic_chars / total
    if ratio > 0.70:
        return "ar", True
    if ratio > 0.20:
        return "mixed", True
    return "en", False


def _cell_to_str(value: object) -> str:
    """Convert any openpyxl cell value to a clean string."""
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKC", text)
    return text


def _is_header_row(row: list[str]) -> bool:
    """
    Heuristic: a row is a header if it has at least 2 non-empty cells,
    and fewer than half the non-empty cells are purely numeric.
    """
    non_empty = [c for c in row if c.strip()]
    if len(non_empty) < 2:
        return False
    numeric_count = sum(1 for c in non_empty if re.fullmatch(r"[-+]?\d+([.,]\d+)?", c.replace(",", "").replace(" ", "")))
    return numeric_count < len(non_empty) / 2


class ExcelParser(BaseParser):
    """
    Parse an Excel file into a DocumentRepresentation.
    Each sheet becomes one DocumentTable; a sheet's text is accumulated into the
    shared all_rows and pages lists.
    """

    async def parse(self, file_bytes: bytes, filename: str) -> dict:
        rep = self._parse_to_representation(file_bytes, filename)
        return {"representation": rep}

    def _parse_to_representation(
        self, file_bytes: bytes, filename: str
    ) -> DocumentRepresentation:
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

        import openpyxl  # type: ignore

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise ValueError(f"Cannot open Excel file: {exc}") from exc

        tables: list[DocumentTable] = []
        all_rows: list[str] = []
        pages: list[str] = []          # one "page" per sheet
        blocks: list[DocumentBlock] = []
        block_counter = 0
        table_counter = 0
        warnings: list[str] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            sheet_rows_raw: list[list[str]] = []

            for row in ws.iter_rows(values_only=True):
                str_row = [_cell_to_str(cell) for cell in row]
                # Skip completely empty rows
                if not any(c for c in str_row):
                    continue
                sheet_rows_raw.append(str_row)

            if not sheet_rows_raw:
                warnings.append(f"Sheet '{sheet_name}': no data found.")
                continue

            # Detect header row
            headers: list[str] = []
            data_rows: list[list[str]] = []
            header_found = False

            for row in sheet_rows_raw:
                if not header_found and _is_header_row(row):
                    headers = row
                    header_found = True
                else:
                    data_rows.append(row)

            if not header_found and sheet_rows_raw:
                # Use first row as header even if it looks numeric
                headers = sheet_rows_raw[0]
                data_rows = sheet_rows_raw[1:]

            tables.append(DocumentTable(
                table_id=f"t{table_counter}",
                page=sheet_idx,
                headers=headers,
                rows=data_rows,
                raw_rows=sheet_rows_raw,
                sheet_name=sheet_name,
            ))
            table_counter += 1

            # Build flat text representation of this sheet
            sheet_lines: list[str] = []
            if headers:
                header_line = " | ".join(h for h in headers if h)
                all_rows.append(header_line)
                sheet_lines.append(header_line)
                blocks.append(DocumentBlock(
                    block_id=f"b{block_counter}",
                    block_type="heading",
                    raw_text=header_line,
                    page=sheet_idx,
                ))
                block_counter += 1

            for row in data_rows:
                row_line = " | ".join(c for c in row if c)
                if not row_line.strip():
                    continue
                all_rows.append(row_line)
                sheet_lines.append(row_line)
                blocks.append(DocumentBlock(
                    block_id=f"b{block_counter}",
                    block_type="text",
                    raw_text=row_line,
                    page=sheet_idx,
                ))
                block_counter += 1

            pages.append("\n".join(sheet_lines))

        wb.close()

        full_text = "\n\n".join(p for p in pages if p)
        lang, has_rtl = _detect_language(full_text)

        return DocumentRepresentation(
            source_format="excel",
            filename=filename,
            full_text=full_text,
            pages=pages,
            blocks=blocks,
            tables=tables,
            all_rows=all_rows,
            coordinates_available=False,
            language_hint=lang,
            has_rtl=has_rtl,
            parse_warnings=warnings,
            page_count=len(wb.sheetnames) if hasattr(wb, "sheetnames") else len(pages),
        )
